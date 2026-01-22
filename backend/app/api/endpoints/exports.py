# -*- coding: utf-8 -*-
"""
Export endpoints for calculation results (PDF, Excel, gbXML).
"""

from datetime import datetime
from io import BytesIO
from typing import Optional
import xml.etree.ElementTree as ET

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse

from app.api.endpoints.auth import get_current_user
from app.api.endpoints.projects import _projects_store

router = APIRouter()


@router.get("/{project_id}/pdf")
async def export_pdf(
    project_id: str,
    current_user: dict = Depends(get_current_user),
):
    """
    Export calculation results as PDF report.
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    project = _projects_store.get(project_id)
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    if project.get("user_id") != current_user["id"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized",
        )

    if "calculation_results" not in project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No calculation results found. Please run a calculation first.",
        )

    results = project["calculation_results"]

    # Create PDF in memory
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
        alignment=TA_CENTER,
    )
    story.append(Paragraph("HVAC Load Calculation Report", title_style))
    story.append(Spacer(1, 0.2 * inch))

    # Project Info
    story.append(Paragraph(f"<b>Project:</b> {project.get('name', 'Untitled')}", styles['Normal']))
    story.append(Paragraph(f"<b>Date:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    if results.get('calculated_at'):
        calc_date = results['calculated_at']
        if isinstance(calc_date, str):
            story.append(Paragraph(f"<b>Calculated:</b> {calc_date}", styles['Normal']))
    story.append(Spacer(1, 0.3 * inch))

    # Building Summary
    story.append(Paragraph("<b>Building Summary</b>", styles['Heading2']))
    summary = results.get('building_summary', {})
    summary_data = [
        ['Metric', 'Cooling', 'Heating'],
        [
            'Peak Load (kW)',
            f"{summary.get('peak_cooling_total', 0) / 1000:.1f}",
            f"{summary.get('peak_heating', 0) / 1000:.1f}"
        ],
        [
            'Load Intensity (W/m²)',
            f"{summary.get('cooling_w_per_m2', 0):.1f}",
            f"{summary.get('heating_w_per_m2', 0):.1f}"
        ],
        [
            'Total Floor Area (m²)',
            f"{summary.get('total_floor_area', 0):.1f}",
            ''
        ],
    ]

    summary_table = Table(summary_data, colWidths=[2.5 * inch, 1.5 * inch, 1.5 * inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.3 * inch))

    # Space Results
    story.append(Paragraph("<b>Space-by-Space Results</b>", styles['Heading2']))
    space_data = [['Space Name', 'Area (m²)', 'Cooling (kW)', 'Heating (kW)', 'Airflow (m³/s)']]

    for space in results.get('space_results', []):
        space_data.append([
            space.get('space_name', ''),
            f"{space.get('floor_area', 0):.1f}",
            f"{space.get('peak_cooling_total', 0) / 1000:.1f}",
            f"{space.get('peak_heating', 0) / 1000:.1f}",
            f"{space.get('supply_airflow_cooling', 0):.1f}",
        ])

    space_table = Table(space_data, colWidths=[2 * inch, 1 * inch, 1 * inch, 1 * inch, 1.2 * inch])
    space_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    story.append(space_table)

    # Build PDF
    doc.build(story)
    buffer.seek(0)

    filename = f"{project.get('name', 'project').replace(' ', '_')}_HVACplus_Report.pdf"

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/{project_id}/excel")
async def export_excel(
    project_id: str,
    current_user: dict = Depends(get_current_user),
):
    """
    Export calculation results as Excel spreadsheet.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    project = _projects_store.get(project_id)
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    if project.get("user_id") != current_user["id"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized",
        )

    if "calculation_results" not in project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No calculation results found. Please run a calculation first.",
        )

    results = project["calculation_results"]

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Summary Sheet
    ws_summary = wb.create_sheet("Building Summary")
    summary = results.get('building_summary', {})

    # Header styling
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_alignment = Alignment(horizontal="center", vertical="center")

    ws_summary['A1'] = 'HVACplus Load Calculation Report'
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A2'] = f"Project: {project.get('name', 'Untitled')}"
    ws_summary['A3'] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    # Building Summary Table
    ws_summary['A5'] = 'Metric'
    ws_summary['B5'] = 'Cooling'
    ws_summary['C5'] = 'Heating'
    for cell in ['A5', 'B5', 'C5']:
        ws_summary[cell].fill = header_fill
        ws_summary[cell].font = header_font
        ws_summary[cell].alignment = center_alignment

    ws_summary['A6'] = 'Peak Load (kW)'
    ws_summary['B6'] = round(summary.get('peak_cooling_total', 0) / 1000, 1)
    ws_summary['C6'] = round(summary.get('peak_heating', 0) / 1000, 1)

    ws_summary['A7'] = 'Load Intensity (W/m²)'
    ws_summary['B7'] = round(summary.get('cooling_w_per_m2', 0), 1)
    ws_summary['C7'] = round(summary.get('heating_w_per_m2', 0), 1)

    ws_summary['A8'] = 'Total Floor Area (m²)'
    ws_summary['B8'] = round(summary.get('total_floor_area', 0), 1)

    # Auto-adjust column widths
    for col in ['A', 'B', 'C']:
        ws_summary.column_dimensions[col].width = 20

    # Space Results Sheet
    ws_spaces = wb.create_sheet("Space Results")

    headers = ['Space Name', 'Floor Area (m²)', 'Cooling Load (kW)', 'Heating Load (kW)',
               'Cooling W/m²', 'Heating W/m²', 'Supply Airflow (m³/s)', 'OA Flow (m³/s)']

    for col_num, header in enumerate(headers, 1):
        cell = ws_spaces.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    # Add space data
    for row_num, space in enumerate(results.get('space_results', []), 2):
        ws_spaces.cell(row=row_num, column=1, value=space.get('space_name', ''))
        ws_spaces.cell(row=row_num, column=2, value=round(space.get('floor_area', 0), 1))
        ws_spaces.cell(row=row_num, column=3, value=round(space.get('peak_cooling_total', 0) / 1000, 1))
        ws_spaces.cell(row=row_num, column=4, value=round(space.get('peak_heating', 0) / 1000, 1))
        ws_spaces.cell(row=row_num, column=5, value=round(space.get('cooling_w_per_m2', 0), 1))
        ws_spaces.cell(row=row_num, column=6, value=round(space.get('heating_w_per_m2', 0), 1))
        ws_spaces.cell(row=row_num, column=7, value=round(space.get('supply_airflow_cooling', 0), 1))
        ws_spaces.cell(row=row_num, column=8, value=round(space.get('outdoor_airflow', 0), 1))

    # Auto-adjust column widths
    for col in range(1, 9):
        ws_spaces.column_dimensions[get_column_letter(col)].width = 18

    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{project.get('name', 'project').replace(' ', '_')}_HVACplus_Results.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/{project_id}/gbxml")
async def export_gbxml(
    project_id: str,
    current_user: dict = Depends(get_current_user),
):
    """
    Export building geometry and loads as gbXML (Green Building XML).
    """
    project = _projects_store.get(project_id)
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    if project.get("user_id") != current_user["id"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized",
        )

    if "calculation_results" not in project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No calculation results found. Please run a calculation first.",
        )

    results = project["calculation_results"]

    # Create gbXML structure
    gbxml = ET.Element("gbXML", attrib={
        "xmlns": "http://www.gbxml.org/schema",
        "version": "6.01",
        "useSIUnitsForResults": "true",
    })

    # Campus element
    campus = ET.SubElement(gbxml, "Campus", id="campus-1")
    ET.SubElement(campus, "Name").text = project.get('name', 'Untitled Project')

    location = ET.SubElement(campus, "Location")
    ET.SubElement(location, "Name").text = project.get('city', 'Unknown')
    ET.SubElement(location, "Latitude").text = str(project.get('latitude', 0.0))
    ET.SubElement(location, "Longitude").text = str(project.get('longitude', 0.0))

    # Building element
    building = ET.SubElement(campus, "Building", id="building-1", buildingType=project.get('building_type', 'Office'))
    ET.SubElement(building, "Name").text = project.get('name', 'Building')
    ET.SubElement(building, "Area", unit="SquareMeters").text = str(round(results.get('building_summary', {}).get('total_floor_area', 0), 1))

    # Spaces
    for idx, space in enumerate(results.get('space_results', []), 1):
        space_elem = ET.SubElement(building, "Space", id=f"space-{idx}")
        ET.SubElement(space_elem, "Name").text = space.get('space_name', f'Space {idx}')
        ET.SubElement(space_elem, "Area", unit="SquareMeters").text = str(round(space.get('floor_area', 0), 1))
        ET.SubElement(space_elem, "Volume", unit="CubicMeters").text = str(round(space.get('volume', 0), 1))

        # Loads
        loads = ET.SubElement(space_elem, "Loads")
        cooling = ET.SubElement(loads, "CoolingLoad", unit="Watts")
        cooling.text = str(round(space.get('peak_cooling_total', 0), 1))

        heating = ET.SubElement(loads, "HeatingLoad", unit="Watts")
        heating.text = str(round(space.get('peak_heating', 0), 1))

        # Airflow
        airflow = ET.SubElement(space_elem, "AirFlow", unit="CubicMetersPerSecond")
        airflow.text = str(round(space.get('supply_airflow_cooling', 0), 1))

    # Convert to string with declaration
    xml_str = '<?xml version="1.0" encoding="UTF-8"?>\n'
    xml_str += ET.tostring(gbxml, encoding='unicode')

    # Format for readability
    import xml.dom.minidom as minidom
    dom = minidom.parseString(xml_str)
    formatted_xml = dom.toprettyxml(indent="  ")

    buffer = BytesIO(formatted_xml.encode('utf-8'))
    buffer.seek(0)

    filename = f"{project.get('name', 'project').replace(' ', '_')}_HVACplus.xml"

    return StreamingResponse(
        buffer,
        media_type="application/xml",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
